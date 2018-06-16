import os, logging, sys, json, csv
sys.path.insert(0, '../')

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from talib.abstract import *

import datetime
import time
import gc

from util.util import getScore, all_day_pct_change_negative, all_day_pct_change_positive, no_doji_or_spinning_buy_india, no_doji_or_spinning_sell_india, scrip_patterns_to_dict
from util.util import is_algo_buy, is_algo_sell
from util.util import get_regressionResult
from util.util import buy_pattern_from_history, buy_all_rule, buy_year_high, buy_year_low, buy_up_trend, buy_down_trend, buy_final, buy_high_indicators, buy_pattern
from util.util import sell_pattern_from_history, sell_all_rule, sell_year_high, sell_year_low, sell_up_trend, sell_down_trend, sell_final, sell_high_indicators, sell_pattern
from util.util import buy_all_filter, buy_all_common, sell_all_filter, sell_all_common
from util.util import buy_all_rule_classifier, sell_all_rule_classifier

connection = MongoClient('localhost', 27017)
db = connection.Nsedata

buyPatternsDict=scrip_patterns_to_dict('../../data-import/nselist/patterns-buy.csv')
sellPatternsDict=scrip_patterns_to_dict('../../data-import/nselist/patterns-sell.csv')

directory = '../../output/final'
logname = '../../output/final' + '/all-result' + time.strftime("%d%m%y-%H%M%S")

newsDict = {}
wb = Workbook()
ws_buyAll = wb.create_sheet("BuyAll")
ws_buyAll.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Avg", "Count"])
ws_buyAllCommon = wb.create_sheet("BuyAllCommon")
ws_buyAllCommon.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Avg", "Count"])
ws_buyAllFilter = wb.create_sheet("BuyAllFilter")
ws_buyAllFilter.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Avg", "Count"])

ws_sellAll = wb.create_sheet("SellAll")
ws_sellAll.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Avg", "Count"])
ws_sellAllCommon = wb.create_sheet("SellAllCommon")
ws_sellAllCommon.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Avg", "Count"])
ws_sellAllFilter = wb.create_sheet("SellAllFilter")
ws_sellAllFilter.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Avg", "Count"])



def saveReports(run_type=None):
    ws_buyAll.append([""])
    ws_buyAllCommon.append([""])
    ws_buyAllFilter.append([""])
        
    ws_sellAll.append([""])
    ws_sellAllCommon.append([""])
    ws_sellAllFilter.append([""])
    

    
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws_buyAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:Y" + str(count))
    tab.tableStyleInfo = style
    ws_buyAll.add_table(tab)
    
    count = 0
    for row in ws_buyAllCommon.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:Y" + str(count))
    tab.tableStyleInfo = style
    ws_buyAllCommon.add_table(tab)
    
    count = 0
    for row in ws_buyAllFilter.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:Y" + str(count))
    tab.tableStyleInfo = style
    ws_buyAllFilter.add_table(tab)
    
    count = 0
    for row in ws_sellAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:Y" + str(count))
    tab.tableStyleInfo = style
    ws_sellAll.add_table(tab)
    
    count = 0
    for row in ws_sellAllCommon.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:Y" + str(count))
    tab.tableStyleInfo = style
    ws_sellAllCommon.add_table(tab)
    
    count = 0
    for row in ws_sellAllFilter.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:Y" + str(count))
    tab.tableStyleInfo = style
    ws_sellAllFilter.add_table(tab)
    
    wb.save(logname + ".xlsx")
      

def result_data(scrip):
    regression_data = db.regressionhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_data = db.classificationhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None and classification_data is not None
        and is_algo_buy(regression_data) and is_algo_buy(classification_data)
    ):
        regressionResult = get_regressionResult(regression_data, scrip, db)
        buyIndiaAvg, result = buy_pattern_from_history(regression_data, regressionResult, None)
        if (buy_all_rule(regression_data, regressionResult, buyIndiaAvg, ws_buyAll)
            or buy_all_rule_classifier(regression_data, regressionResult, buyIndiaAvg, ws_buyAll)):
            buy_all_filter(regression_data, regressionResult, ws_buyAllFilter)
            buy_all_common(regression_data, classification_data, regressionResult, ws_buyAllCommon)
        
        regressionResult = get_regressionResult(classification_data, scrip, db)
        buyIndiaAvg, result = buy_pattern_from_history(classification_data, regressionResult, None)
        if (buy_all_rule(regression_data, regressionResult, buyIndiaAvg, ws_buyAll)
            or buy_all_rule_classifier(regression_data, regressionResult, buyIndiaAvg, ws_buyAll)):
            buy_all_filter(classification_data, regressionResult, ws_buyAllFilter)
            buy_all_common(regression_data, classification_data, regressionResult, ws_buyAllCommon)
                 
    regression_data = db.regressionlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_data = db.classificationlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None and classification_data is not None
        and is_algo_sell(regression_data) and is_algo_sell(classification_data)
    ):
        regressionResult = get_regressionResult(regression_data, scrip, db)
        sellIndiaAvg, result = sell_pattern_from_history(regression_data, regressionResult, None)
        if (sell_all_rule(regression_data, regressionResult, sellIndiaAvg, ws_sellAll)
            or sell_all_rule_classifier(regression_data, regressionResult, sellIndiaAvg, ws_sellAll)):
            sell_all_filter(regression_data, regressionResult, ws_sellAllFilter)
            sell_all_common(regression_data, classification_data, regressionResult, ws_sellAllCommon)
        
        regressionResult = get_regressionResult(classification_data, scrip, db)
        sellIndiaAvg, result = sell_pattern_from_history(classification_data, regressionResult, None)
        if (sell_all_rule(regression_data, regressionResult, sellIndiaAvg, ws_sellAll)
            or sell_all_rule_classifier(regression_data, regressionResult, sellIndiaAvg, ws_sellAll)):
            sell_all_filter(classification_data, regressionResult, ws_sellAllFilter)
            sell_all_common(regression_data, classification_data, regressionResult, ws_sellAllCommon)
              
                                    
                          
def calculateParallel(threads=2, futures=None):
    pool = ThreadPool(threads)
    scrips = []
    for data in db.scrip.find({'futures':futures}):
        scrips.append(data['scrip'].replace('&','').replace('-','_'))
    scrips.sort()
    pool.map(result_data, scrips)       
                     
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1, sys.argv[1])
    connection.close()
    saveReports(sys.argv[1])