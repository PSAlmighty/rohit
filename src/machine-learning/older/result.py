import os, logging, sys, json, csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from talib.abstract import *

import datetime
import time
import gc

connection = MongoClient('localhost', 27017)
db = connection.Nsedata

directory = '../../output/final'
logname = '../../output/final' + '/results' + time.strftime("%d%m%y-%H%M%S")

newsDict = {}
wb = Workbook()
ws = wb.active
ws.append(["scrip", "timestamps", "summary", "Link"])
ws_buyFilter = wb.create_sheet("BuyFilter")
ws_buyFilter.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sellFilter = wb.create_sheet("SellFilter")
ws_sellFilter.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buyNews = wb.create_sheet("BuyNews")
ws_buyNews.append(["timestamps", "summary", "Link"])
ws_sellNews = wb.create_sheet("SellNews")
ws_sellNews.append(["timestamps", "summary", "Link"])
ws_buy = wb.create_sheet("Buy")
ws_buy.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sell = wb.create_sheet("Sell")
ws_sell.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buyAll = wb.create_sheet("BuyAll")
ws_buyAll.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sellAll = wb.create_sheet("SellAll")
ws_sellAll.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buyFinal = wb.create_sheet("BuyFinal")
ws_buyFinal.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sellFinal = wb.create_sheet("SellFinal")
ws_sellFinal.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buyFinal1 = wb.create_sheet("BuyFinal1")
ws_buyFinal1.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sellFinal1 = wb.create_sheet("SellFinal1")
ws_sellFinal1.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])


def saveDailyNews():
    for newslink,newsValue in newsDict.items():
        ws.append([newsValue['scrip'], newsValue['newstime'], newsValue['newssummary'], newslink])

def saveReports(run_type=None):
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:D" + str(count))
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    count = 0
    for row in ws_buyFilter.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyFilter.add_table(tab)
    
    count = 0
    for row in ws_sellFilter.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellFilter.add_table(tab)
    
    count = 0
    for row in ws_buyNews.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:C" + str(count))
    tab.tableStyleInfo = style
    ws_buyNews.add_table(tab)
    
    count = 0
    for row in ws_sellNews.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:C" + str(count))
    tab.tableStyleInfo = style
    ws_sellNews.add_table(tab)
    
    count = 0
    for row in ws_buy.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buy.add_table(tab)
    
    count = 0
    for row in ws_sell.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sell.add_table(tab)
    
    count = 0
    for row in ws_buyAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyAll.add_table(tab)
    
    count = 0
    for row in ws_sellAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellAll.add_table(tab)
    
    count = 0
    for row in ws_buyFinal.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyFinal.add_table(tab)
    
    count = 0
    for row in ws_sellFinal.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellFinal.add_table(tab)
    
    count = 0
    for row in ws_buyFinal1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyFinal1.add_table(tab)
    
    count = 0
    for row in ws_sellFinal1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellFinal1.add_table(tab)
      
    if(run_type == 'broker'):
        wb.save(logname + "broker_buy.xlsx")
    elif(run_type == 'result'):
        wb.save(logname + "result.xlsx")
    elif(run_type == 'result_declared'):
        wb.save(logname + "result_declared.xlsx")       
    else:
        wb.save(logname + ".xlsx")

def buy_News(scrip):
    scrip_newsList = db.news.find_one({'scrip':scrip})
    ws_buyNews.append([scrip])
    ws_buyNews.append(["#####################"])
    if(scrip_newsList is None):
        print('Missing news for ', scrip)
        return
    
    for scrip_news in scrip_newsList['news']:
        ws_buyNews.append([scrip_news['timestamp'], scrip_news['summary'], scrip_news['link']])
    ws_buyNews.append([" "])
    
def sell_News(scrip):
    scrip_newsList = db.news.find_one({'scrip':scrip})
    ws_sellNews.append([scrip])
    ws_sellNews.append(["#####################"])
    if(scrip_newsList is None):
        print('Missing news for ', scrip)
        return
    for scrip_news in scrip_newsList['news']:
        ws_sellNews.append([scrip_news['timestamp'], scrip_news['summary'], scrip_news['link']])
    ws_sellNews.append([" "])  

def result_data(scrip):
    classification_data = db.classification.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    regression_data = db.regression.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    
    if(classification_data is None or regression_data is None):
        print('Missing or very less Data for ', scrip)
        return
    
    #Buy Indicators
    regressionResult = [ ]
    regressionResult.append(regression_data['futures'])
    regressionResult.append(regression_data['trainSize'])
    regressionResult.append(regression_data['buyIndia'])
    regressionResult.append(regression_data['sellIndia'])
    regressionResult.append(regression_data['scrip'])
    regressionResult.append(regression_data['forecast_day_VOL_change'])
    regressionResult.append(regression_data['forecast_day_PCT_change'])
    regressionResult.append(regression_data['forecast_day_PCT2_change'])
    regressionResult.append(regression_data['forecast_day_PCT3_change'])
    regressionResult.append(regression_data['forecast_day_PCT4_change'])
    regressionResult.append(regression_data['forecast_day_PCT5_change'])
    regressionResult.append(regression_data['forecast_day_PCT7_change'])
    regressionResult.append(regression_data['forecast_day_PCT10_change'])
    regressionResult.append(regression_data['PCT_day_change'])
    regressionResult.append(regression_data['score'])
    regressionResult.append(regression_data['randomForestValue'])
    regressionResult.append(regression_data['randomForestAccuracy'])
    regressionResult.append(regression_data['mlpValue'])
    regressionResult.append(regression_data['mlpAccuracy'])
    regressionResult.append(regression_data['baggingValue'])
    regressionResult.append(regression_data['baggingAccuracy'])
    regressionResult.append(regression_data['adaBoostValue'])
    regressionResult.append(regression_data['adaBoostAccuracy'])
    regressionResult.append(regression_data['kNeighboursValue'])
    regressionResult.append(regression_data['kNeighboursAccuracy'])
    regressionResult.append(regression_data['gradientBoostingValue'])
    regressionResult.append(regression_data['gradientBoostingAccuracy'])
    regressionResult.append(regression_data['trend'])
    regressionResult.append(regression_data['yearHighChange'])
    regressionResult.append(regression_data['yearLowChange'])
    if(classification_data['kNeighboursValue'] >= 0 and regression_data['kNeighboursValue'] > .5):
        ws_buy.append(regressionResult)
        if(regression_data['kNeighboursValue'] >= .5 
           and regression_data['mlpValue'] >= .5  
           and (float(regression_data['mlpValue']) + float(regression_data['randomForestValue'])) >= 1
           and regression_data['baggingValue'] >= 0 
           and regression_data['yearHighChange'] <= -10):
            ws_buyFilter.append(regressionResult) 
        if(regression_data['mlpValue'] > .5 and regression_data['kNeighboursValue'] > .5 and regression_data['forecast_day_PCT7_change'] < -1 and regression_data['forecast_day_PCT10_change'] < 1 and 5 > regression_data['PCT_day_change'] > -.5 and 5 > regression_data['forecast_day_PCT_change'] > 0):
            ws_buyFinal.append(regressionResult) 
        elif(regression_data['mlpValue'] > 0 and regression_data['kNeighboursValue'] > .5 and 5 > regression_data['PCT_day_change'] > -.5 and 5 > regression_data['forecast_day_PCT_change'] > -1):
            ws_buyFinal1.append(regressionResult)             
    if(regression_data['kNeighboursValue'] > 1 and regression_data['mlpValue'] > .5):
        ws_buyAll.append(regressionResult)  
    if(regression_data['kNeighboursValue'] > 0): 
        buy_News(scrip)     
        
    #Sell Indicators
    regressionResult = [ ]
    regressionResult.append(classification_data['futures'])
    regressionResult.append(classification_data['trainSize'])
    regressionResult.append(classification_data['buyIndia'])
    regressionResult.append(classification_data['sellIndia'])
    regressionResult.append(classification_data['scrip'])
    regressionResult.append(classification_data['forecast_day_VOL_change'])
    regressionResult.append(classification_data['forecast_day_PCT_change'])
    regressionResult.append(classification_data['forecast_day_PCT2_change'])
    regressionResult.append(classification_data['forecast_day_PCT3_change'])
    regressionResult.append(classification_data['forecast_day_PCT4_change'])
    regressionResult.append(classification_data['forecast_day_PCT5_change'])
    regressionResult.append(classification_data['forecast_day_PCT7_change'])
    regressionResult.append(classification_data['forecast_day_PCT10_change'])
    regressionResult.append(classification_data['PCT_day_change'])
    regressionResult.append(classification_data['score'])
    regressionResult.append(classification_data['randomForestValue'])
    regressionResult.append(classification_data['randomForestAccuracy'])
    regressionResult.append(classification_data['mlpValue'])
    regressionResult.append(classification_data['mlpAccuracy'])
    regressionResult.append(classification_data['baggingValue'])
    regressionResult.append(classification_data['baggingAccuracy'])
    regressionResult.append(classification_data['adaBoostValue'])
    regressionResult.append(classification_data['adaBoostAccuracy'])
    regressionResult.append(classification_data['kNeighboursValue'])
    regressionResult.append(classification_data['kNeighboursAccuracy'])
    regressionResult.append(classification_data['gradientBoostingValue'])
    regressionResult.append(classification_data['gradientBoostingAccuracy'])
    regressionResult.append(classification_data['trend'])
    regressionResult.append(classification_data['yearHighChange'])
    regressionResult.append(classification_data['yearLowChange'])
    if(classification_data['kNeighboursValue'] < 0 and regression_data['kNeighboursValue'] <= .5):            
        ws_sell.append(regressionResult)
        if(float(classification_data['yearHighChange']) > -30):
            ws_sellFilter.append(regressionResult)    
    if(classification_data['kNeighboursValue'] < 0 and classification_data['mlpValue'] <= 0):            
        ws_sellAll.append(regressionResult) 
    if(classification_data['kNeighboursValue'] < 0 and classification_data['mlpValue'] <= 0 and classification_data['forecast_day_PCT7_change'] > 1 and classification_data['forecast_day_PCT10_change'] > -1 and -5 < classification_data['PCT_day_change'] < .5 and -5 < classification_data['forecast_day_PCT_change'] < 0):
        ws_sellFinal.append(regressionResult)
    elif(classification_data['kNeighboursValue'] < 0 and classification_data['mlpValue'] <= 0 and -5 < classification_data['PCT_day_change'] < .5 and -5 < classification_data['forecast_day_PCT_change'] < 1):
        ws_sellFinal1.append(regressionResult)    
    if(classification_data['kNeighboursValue'] < 0):
        sell_News(scrip)  
        
    start_date = (datetime.datetime.now() - datetime.timedelta(hours=0))
    start_date = datetime.datetime(start_date.year, start_date.month, start_date.day, start_date.hour) 
    end_date = (datetime.datetime.now() - datetime.timedelta(hours=18))
    end_date = datetime.datetime(end_date.year, end_date.month, end_date.day, end_date.hour)
    
    scrip_newsList = db.news.find_one({'scrip':scrip})
    if(scrip_newsList is None):
        return
    for news in scrip_newsList['news']:
        newstime = news['timestamp']
        newssummary = news['summary']
        newslink = news['link']
        try:
            news_time = datetime.datetime.strptime(newstime, "%H:%M:%S %d-%m-%Y")
            if start_date > news_time > end_date: 
                if newslink in newsDict:
                    newsDict[newslink]['scrip'] = newsDict[newslink]['scrip'] + ',' + scrip
                else:
                    newsValue = {}
                    newsValue['newssummary'] = newssummary
                    newsValue['newstime'] = newstime
                    newsValue['scrip'] = scrip
                    newsDict[newslink] = newsValue
                
                
        except:
            pass
                              
def calculateParallel(threads=2, run_type=None, futures=None):
    pool = ThreadPool(threads)
    if(run_type == 'broker'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_broker_buy.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)
    elif(run_type == 'result'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_result.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)  
    elif(run_type == 'result_declared'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_result_declared.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)               
    else:
        scrips = []
        for data in db.scrip.find({'futures':futures}):
            scrips.append(data['scrip'].replace('&','').replace('-','_'))
        scrips.sort()
        pool.map(result_data, scrips)
                     
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1, sys.argv[1], sys.argv[2])
    connection.close()
    saveDailyNews()
    saveReports(sys.argv[1])