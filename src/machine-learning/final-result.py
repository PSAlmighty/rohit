import os, logging, sys, json, csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from talib.abstract import *

import datetime
import time
import gc

connection = MongoClient('localhost', 27017)
db = connection.Nsedata

directory = '../../output/final'
logname = '../../output/final' + '/final-result' + time.strftime("%d%m%y-%H%M%S")

newsDict = {}
wb = Workbook()
ws_buyAll = wb.create_sheet("BuyAll")
ws_buyAll.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_sellAll = wb.create_sheet("SellAll")
ws_sellAll.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_buyYearHigh = wb.create_sheet("buyYearHigh")
ws_buyYearHigh.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_sellYearHigh = wb.create_sheet("sellYearHigh")
ws_sellYearHigh.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_buyYearLow = wb.create_sheet("buyYearLow")
ws_buyYearLow.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_sellYearLow = wb.create_sheet("sellYearLow")
ws_sellYearLow.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_buyFinal = wb.create_sheet("BuyFinal")
ws_buyFinal.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_sellFinal = wb.create_sheet("SellFinal")
ws_sellFinal.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_buyFinal1 = wb.create_sheet("BuyFinal1")
ws_buyFinal1.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_sellFinal1 = wb.create_sheet("SellFinal1")
ws_sellFinal1.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_buyPattern = wb.create_sheet("BuyPattern")
ws_buyPattern.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_sellPattern = wb.create_sheet("SellPattern")
ws_sellPattern.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_buyPattern1 = wb.create_sheet("BuyPattern1")
ws_buyPattern1.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_sellPattern1 = wb.create_sheet("SellPattern1")
ws_sellPattern1.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_buyPattern2 = wb.create_sheet("buyPattern2")
ws_buyPattern2.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])
ws_sellPattern2 = wb.create_sheet("sellPattern2")
ws_sellPattern2.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score", "MLP", "accuracy", "KNeighbors", "accuracy", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment"])

def saveReports(run_type=None):
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws_buyAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyAll.add_table(tab)
    
    count = 0
    for row in ws_sellAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellAll.add_table(tab)
    
    count = 0
    for row in ws_buyYearHigh.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyYearHigh.add_table(tab)
    
    count = 0
    for row in ws_sellYearHigh.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellYearHigh.add_table(tab)
    
    count = 0
    for row in ws_buyYearLow.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyYearLow.add_table(tab)
    
    count = 0
    for row in ws_sellYearLow.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellYearLow.add_table(tab)
    
    count = 0
    for row in ws_buyFinal.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyFinal.add_table(tab)
    
    count = 0
    for row in ws_sellFinal.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellFinal.add_table(tab)
    
    count = 0
    for row in ws_buyFinal1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyFinal1.add_table(tab)
    
    count = 0
    for row in ws_sellFinal1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellFinal1.add_table(tab)
    
    count = 0
    for row in ws_buyPattern.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyPattern.add_table(tab)
    
    count = 0
    for row in ws_sellPattern.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellPattern.add_table(tab)
      
    count = 0
    for row in ws_buyPattern1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyPattern1.add_table(tab)
    
    count = 0
    for row in ws_sellPattern1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellPattern1.add_table(tab)
    
    count = 0
    for row in ws_buyPattern2.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyPattern2.add_table(tab)
    
    count = 0
    for row in ws_sellPattern2.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellPattern2.add_table(tab)
      
    if(run_type == 'broker'):
        wb.save(logname + "broker_buy.xlsx")
    elif(run_type == 'result'):
        wb.save(logname + "result.xlsx")
    elif(run_type == 'result_declared'):
        wb.save(logname + "result_declared.xlsx")       
    else:
        wb.save(logname + ".xlsx")

def buy_News(scrip):
    scrip_newsList = db.news.find_one({'scrip':scrip})
    ws_buyNews.append([scrip])
    ws_buyNews.append(["#####################"])
    if(scrip_newsList is None):
        print('Missing news for ', scrip)
        return
    
    for scrip_news in scrip_newsList['news']:
        ws_buyNews.append([scrip_news['timestamp'], scrip_news['summary'], scrip_news['link']])
    ws_buyNews.append([" "])
    
def sell_News(scrip):
    scrip_newsList = db.news.find_one({'scrip':scrip})
    ws_sellNews.append([scrip])
    ws_sellNews.append(["#####################"])
    if(scrip_newsList is None):
        print('Missing news for ', scrip)
        return
    for scrip_news in scrip_newsList['news']:
        ws_sellNews.append([scrip_news['timestamp'], scrip_news['summary'], scrip_news['link']])
    ws_sellNews.append([" "])  

def result_data(scrip):
    resultDeclared = ""
    resultDate = ""
    resultSentiment = ""
    result_data = db.scrip_result.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(result_data is not None):
        resultDate = result_data['result_date'].strip()
        resultSentiment = result_data['result_sentiment']
        start_date = (datetime.datetime.now() - datetime.timedelta(hours=0))
        start_date = datetime.datetime(start_date.year, start_date.month, start_date.day, start_date.hour)
        result_time = datetime.datetime.strptime(resultDate, "%Y-%m-%d")
        if result_time < start_date: 
            resultDeclared = resultDate 
            resultDate = ""
        
    regression_data = db.regressionhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_data = db.classificationhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None and regression_data is not None):
        regressionResult = [ ]
        regressionResult.append(regression_data['futures'])
        regressionResult.append(regression_data['trainSize'])
        regressionResult.append(regression_data['buyIndia'])
        regressionResult.append(regression_data['sellIndia'])
        regressionResult.append(regression_data['scrip'])
        regressionResult.append(regression_data['forecast_day_VOL_change'])
        regressionResult.append(regression_data['forecast_day_PCT_change'])
        regressionResult.append(regression_data['forecast_day_PCT2_change'])
        regressionResult.append(regression_data['forecast_day_PCT3_change'])
        regressionResult.append(regression_data['forecast_day_PCT4_change'])
        regressionResult.append(regression_data['forecast_day_PCT5_change'])
        regressionResult.append(regression_data['forecast_day_PCT7_change'])
        regressionResult.append(regression_data['forecast_day_PCT10_change'])
        regressionResult.append(regression_data['PCT_day_change'])
        regressionResult.append(regression_data['score'])
        regressionResult.append(regression_data['mlpValue'])
        regressionResult.append(regression_data['mlpAccuracy'])
        regressionResult.append(regression_data['kNeighboursValue'])
        regressionResult.append(regression_data['kNeighboursAccuracy'])
        regressionResult.append(regression_data['trend'])
        regressionResult.append(regression_data['yearHighChange'])
        regressionResult.append(regression_data['yearLowChange'])
        regressionResult.append(resultDate)
        regressionResult.append(resultDeclared)
        regressionResult.append(resultSentiment)
        classificationResult = [ ]
        classificationResult.append(classification_data['futures'])
        classificationResult.append(classification_data['trainSize'])
        classificationResult.append(classification_data['buyIndia'])
        classificationResult.append(classification_data['sellIndia'])
        classificationResult.append(classification_data['scrip'])
        classificationResult.append(classification_data['forecast_day_VOL_change'])
        classificationResult.append(classification_data['forecast_day_PCT_change'])
        classificationResult.append(classification_data['forecast_day_PCT2_change'])
        classificationResult.append(classification_data['forecast_day_PCT3_change'])
        classificationResult.append(classification_data['forecast_day_PCT4_change'])
        classificationResult.append(classification_data['forecast_day_PCT5_change'])
        classificationResult.append(classification_data['forecast_day_PCT7_change'])
        classificationResult.append(classification_data['forecast_day_PCT10_change'])
        classificationResult.append(classification_data['PCT_day_change'])
        classificationResult.append(classification_data['score'])
        classificationResult.append(classification_data['mlpValue'])
        classificationResult.append(classification_data['mlpAccuracy'])
        classificationResult.append(classification_data['kNeighboursValue'])
        classificationResult.append(classification_data['kNeighboursAccuracy'])
        classificationResult.append(classification_data['trend'])
        classificationResult.append(classification_data['yearHighChange'])
        classificationResult.append(classification_data['yearLowChange'])
        classificationResult.append(resultDate)
        classificationResult.append(resultDeclared)
        classificationResult.append(resultSentiment)
        score = ''
        if(regression_data['score'] == '10' or regression_data['score'] == '1-1'):
            score = 'up'
        if(((regression_data['mlpValue'] >= 1 and regression_data['kNeighboursValue'] >= 0.5) or (regression_data['mlpValue'] >= 0.5 and regression_data['kNeighboursValue'] >= 1)) 
            and ((classification_data['mlpValue'] >= 1 and classification_data['kNeighboursValue'] >= 0) or (classification_data['mlpValue'] >= 1 and classification_data['kNeighboursValue'] >= 1))
            and 'P@[' not in str(regression_data['sellIndia'])):
            ws_buyAll.append(regressionResult)
            ws_buyAll.append(classificationResult)
            if(-5 < regression_data['yearHighChange'] < -1 and regression_data['forecast_day_PCT5_change'] <= 5):
                ws_buyYearHigh.append(regressionResult)
                ws_buyYearHigh.append(classificationResult)
            elif(-20 < regression_data['yearHighChange'] < -1 and 0 < regression_data['PCT_day_change'] < 3 and (score == 'up'  or regression_data['forecast_day_PCT_change'] > 0)):
                ws_buyYearHigh.append(regressionResult)
                ws_buyYearHigh.append(classificationResult)
                
            if(1 < regression_data['yearLowChange'] < 10 and 0 < regression_data['PCT_day_change'] < 3 and score == 'up'):
                ws_buyYearLow.append(regressionResult)
                ws_buyYearLow.append(classificationResult)    
                
            if(2 > regression_data['PCT_day_change'] > 0 and str(regression_data['sellIndia']) == '' and -95 < regression_data['yearHighChange'] < -15
                and regression_data['forecast_day_PCT5_change'] <= 1 and regression_data['forecast_day_PCT7_change'] <= 1 and regression_data['forecast_day_PCT10_change'] <= 1):
                ws_buyFinal.append(regressionResult)
                ws_buyFinal.append(classificationResult)
            elif(2 > regression_data['PCT_day_change'] > 0
                and regression_data['forecast_day_PCT5_change'] <= 1 and regression_data['forecast_day_PCT7_change'] <= -1 and regression_data['forecast_day_PCT10_change'] <= -5):
                ws_buyFinal1.append(regressionResult)
                ws_buyFinal1.append(classificationResult)
         
            if(regression_data['PCT_day_change'] < 4 and regression_data['yearLowChange'] > 5):
                if(('MARUBOZU' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and regression_data['forecast_day_PCT10_change'] <= 1)
                   or ('HAMMER' in str(regression_data['buyIndia']) and regression_data['PCT_day_change'] > 0)
                   #or 'ENGULFING' in str(regression_data['buyIndia'])
                   #or 'PIERCING' in str(regression_data['buyIndia'])
                   or ('MORNINGSTAR' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and regression_data['forecast_day_PCT10_change'] <= 1)
                   #or ':DOJISTAR' in str(regression_data['buyIndia'])
                   #or 'MORNINGDOJISTAR' in str(regression_data['buyIndia'])
                   or 'ABANDONEDBABY' in str(regression_data['buyIndia'])
                   or 'COUNTERATTACK' in str(regression_data['buyIndia'])
                   or 'KICKING' in str(regression_data['buyIndia'])
                   or 'BREAKAWAY' in str(regression_data['buyIndia'])
                   #or 'TRISTAR' in str(regression_data['buyIndia'])
                   #or '3WHITESOLDIERS' in str(regression_data['buyIndia'])
                   #or '3INSIDE' in str(regression_data['buyIndia'])
                   ):
                    ws_buyPattern.append(regressionResult)
                    ws_buyPattern.append(classificationResult)       
                elif(('CCI:BOP' in str(regression_data['buyIndia']) and 'BELTHOLD' in str(regression_data['buyIndia']))
                   or ('AROON:BOP' in str(regression_data['buyIndia']) and 'BELTHOLD' in str(regression_data['buyIndia']) and 'ENGULFING' in str(regression_data['buyIndia']))
                   or ('BELTHOLD' == str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and score == 'up')
                   or ('3OUTSIDE' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and score == 'up')
                   or ('HARAMI' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and score == 'up')
                   or (regression_data['yearHighChange'] <= -35 and 'HARAMI' in str(regression_data['buyIndia']) and 'SHORTLINE' in str(regression_data['buyIndia']) and regression_data['PCT_day_change'] > 0)
                   or ('DOJI' in str(regression_data['buyIndia']) and 'GRAVESTONEDOJI' in str(regression_data['buyIndia']) and 'LONGLEGGEDDOJI' in str(regression_data['buyIndia']) and regression_data['PCT_day_change'] > 0)
                   or ('P@[,HIKKAKE]' == str(regression_data['buyIndia']) and regression_data['PCT_day_change'] < 0)
                   #or (regression_data['yearHighChange'] <= -35 and 'BELTHOLD' in str(regression_data['buyIndia']) and 'LONGLINE' in str(regression_data['buyIndia']))
                   #or (regression_data['yearHighChange'] <= -35 and ',CCI:BOP' in str(regression_data['buyIndia']) and 'LONGLINE' in str(regression_data['buyIndia']))
                   ):
                    ws_buyPattern1.append(regressionResult)
                    ws_buyPattern1.append(classificationResult)
                elif((('MARUBOZU' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and regression_data['forecast_day_PCT10_change'] <= 1)
                   or ('HAMMER' in str(regression_data['buyIndia']) and regression_data['PCT_day_change'] > 0)
                   or 'ENGULFING' in str(regression_data['buyIndia'])
                   or 'PIERCING' in str(regression_data['buyIndia'])
                   or ('MORNINGSTAR' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and regression_data['forecast_day_PCT10_change'] <= 1)
                   or ':DOJISTAR' in str(regression_data['buyIndia'])
                   or 'MORNINGDOJISTAR' in str(regression_data['buyIndia'])
                   or 'ABANDONEDBABY' in str(regression_data['buyIndia'])
                   or 'COUNTERATTACK' in str(regression_data['buyIndia'])
                   or 'KICKING' in str(regression_data['buyIndia'])
                   or 'BREAKAWAY' in str(regression_data['buyIndia'])
                   or 'TRISTAR' in str(regression_data['buyIndia'])
                   or '3WHITESOLDIERS' in str(regression_data['buyIndia'])
                   or '3INSIDE' in str(regression_data['buyIndia'])
                   ) and (regression_data['forecast_day_PCT5_change'] <= -5) and (regression_data['forecast_day_PCT10_change'] <= -5)): 
                    ws_buyPattern2.append(regressionResult)
                    ws_buyPattern2.append(classificationResult)                    
    
        
    regression_data = db.regressionlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_data = db.classificationlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None or regression_data is not None):
        regressionResult = [ ]
        regressionResult.append(regression_data['futures'])
        regressionResult.append(regression_data['trainSize'])
        regressionResult.append(regression_data['buyIndia'])
        regressionResult.append(regression_data['sellIndia'])
        regressionResult.append(regression_data['scrip'])
        regressionResult.append(regression_data['forecast_day_VOL_change'])
        regressionResult.append(regression_data['forecast_day_PCT_change'])
        regressionResult.append(regression_data['forecast_day_PCT2_change'])
        regressionResult.append(regression_data['forecast_day_PCT3_change'])
        regressionResult.append(regression_data['forecast_day_PCT4_change'])
        regressionResult.append(regression_data['forecast_day_PCT5_change'])
        regressionResult.append(regression_data['forecast_day_PCT7_change'])
        regressionResult.append(regression_data['forecast_day_PCT10_change'])
        regressionResult.append(regression_data['PCT_day_change'])
        regressionResult.append(regression_data['score'])
        regressionResult.append(regression_data['mlpValue'])
        regressionResult.append(regression_data['mlpAccuracy'])
        regressionResult.append(regression_data['kNeighboursValue'])
        regressionResult.append(regression_data['kNeighboursAccuracy'])
        regressionResult.append(regression_data['trend'])
        regressionResult.append(regression_data['yearHighChange'])
        regressionResult.append(regression_data['yearLowChange'])
        regressionResult.append(resultDate)
        regressionResult.append(resultDeclared)
        regressionResult.append(resultSentiment)
        classificationResult = [ ]
        classificationResult.append(classification_data['futures'])
        classificationResult.append(classification_data['trainSize'])
        classificationResult.append(classification_data['buyIndia'])
        classificationResult.append(classification_data['sellIndia'])
        classificationResult.append(classification_data['scrip'])
        classificationResult.append(classification_data['forecast_day_VOL_change'])
        classificationResult.append(classification_data['forecast_day_PCT_change'])
        classificationResult.append(classification_data['forecast_day_PCT2_change'])
        classificationResult.append(classification_data['forecast_day_PCT3_change'])
        classificationResult.append(classification_data['forecast_day_PCT4_change'])
        classificationResult.append(classification_data['forecast_day_PCT5_change'])
        classificationResult.append(classification_data['forecast_day_PCT7_change'])
        classificationResult.append(classification_data['forecast_day_PCT10_change'])
        classificationResult.append(classification_data['PCT_day_change'])
        classificationResult.append(classification_data['score'])
        classificationResult.append(classification_data['mlpValue'])
        classificationResult.append(classification_data['mlpAccuracy'])
        classificationResult.append(classification_data['kNeighboursValue'])
        classificationResult.append(classification_data['kNeighboursAccuracy'])
        classificationResult.append(classification_data['trend'])
        classificationResult.append(classification_data['yearHighChange'])
        classificationResult.append(classification_data['yearLowChange'])
        classificationResult.append(resultDate)
        classificationResult.append(resultDeclared)
        classificationResult.append(resultSentiment)
        score = ''
        if(regression_data['score'] == '1-1' or regression_data['score'] == '0-1'):
            score = 'down'
        if(((regression_data['mlpValue'] <= -1 and regression_data['kNeighboursValue'] <= -0.5) or (regression_data['mlpValue'] <= -0.5 and regression_data['kNeighboursValue'] <= -1))
            and ((classification_data['mlpValue'] <= -1 and classification_data['kNeighboursValue'] <= 0) or (classification_data['mlpValue'] <= -1 and classification_data['kNeighboursValue'] <= -1))
            and 'P@[' not in str(regression_data['buyIndia'])):
            ws_sellAll.append(regressionResult)
            ws_sellAll.append(classificationResult)
            if(-10 < regression_data['yearHighChange'] < -2 and -2 < regression_data['PCT_day_change'] < 0 and (score == 'down'  or regression_data['forecast_day_PCT_change'] < 0)):
                ws_sellYearHigh.append(regressionResult)
                ws_sellYearHigh.append(classificationResult)
                
            if(1 < regression_data['yearLowChange'] < 15 and -2 < regression_data['PCT_day_change'] < 0 and regression_data['forecast_day_PCT_change'] < 0):
                ws_sellYearLow.append(regressionResult)
                ws_sellYearLow.append(classificationResult)    
            
            if(-2 < regression_data['PCT_day_change'] < 0 and str(regression_data['buyIndia']) == '' and -95 < regression_data['yearHighChange'] < -20
                and regression_data['forecast_day_PCT5_change'] >= 1 and regression_data['forecast_day_PCT7_change'] >= 1 and regression_data['forecast_day_PCT10_change'] >= 1):
                ws_sellFinal.append(regressionResult)
                ws_sellFinal.append(classificationResult)
            elif(-2 < regression_data['PCT_day_change'] < 0 
                and regression_data['forecast_day_PCT5_change'] >= -1 and regression_data['forecast_day_PCT7_change'] >= 1 and regression_data['forecast_day_PCT10_change'] >= 5):
                ws_sellFinal1.append(regressionResult)
                ws_sellFinal1.append(classificationResult)  
                               
            if(-4 < regression_data['PCT_day_change'] and regression_data['yearHighChange'] < -5):
                if(('HANGINGMAN' in str(regression_data['sellIndia'])
                   #or 'MARUBOZU' in str(regression_data['sellIndia'])
                   #or 'ENGULFING' in str(regression_data['sellIndia'])
                   or 'EVENINGSTAR' in str(regression_data['sellIndia'])
                   #or ':DOJISTAR' in str(regression_data['sellIndia'])
                   #or 'EVENINGDOJISTAR' in str(regression_data['sellIndia'])
                   or 'ABANDONEDBABY' in str(regression_data['sellIndia'])
                   or 'COUNTERATTACK' in str(regression_data['sellIndia'])
                   or 'KICKING' in str(regression_data['sellIndia'])
                   or 'BREAKAWAY' in str(regression_data['sellIndia'])
                   #or 'TRISTAR' in str(regression_data['sellIndia'])
                   or ('SHOOTINGSTAR' in str(regression_data['sellIndia']) and regression_data['PCT_day_change'] < 0)
                   or 'DARKCLOUDCOVER' in str(regression_data['sellIndia'])
                   #or '3INSIDE' in str(regression_data['sellIndia'])
                   #or '3OUTSIDE' in str(regression_data['sellIndia'])
                   #or '2CROWS' in str(regression_data['sellIndia'])
                   #or '3BLACKCROWS' in str(regression_data['sellIndia'])
                   ) and (regression_data['forecast_day_PCT5_change'] >= 0)):
                    ws_sellPattern.append(regressionResult)
                    ws_sellPattern.append(classificationResult) 
                elif((('HARAMI' in str(regression_data['sellIndia']) and regression_data['forecast_day_PCT5_change'] >= 0 and score == 'down')
                   or ('ENGULFING' in str(regression_data['sellIndia']) and 'LONGLINE' in str(regression_data['sellIndia']) and score == 'down')
                   ) and regression_data['yearHighChange'] < -5):
                    ws_sellPattern1.append(regressionResult)
                    ws_sellPattern1.append(classificationResult)
                elif(('HANGINGMAN' in str(regression_data['sellIndia'])
                   or 'MARUBOZU' in str(regression_data['sellIndia'])
                   or 'ENGULFING' in str(regression_data['sellIndia'])
                   or 'EVENINGSTAR' in str(regression_data['sellIndia'])
                   #or ':DOJISTAR' in str(regression_data['sellIndia'])
                   #or 'EVENINGDOJISTAR' in str(regression_data['sellIndia'])
                   or 'ABANDONEDBABY' in str(regression_data['sellIndia'])
                   or 'COUNTERATTACK' in str(regression_data['sellIndia'])
                   or 'KICKING' in str(regression_data['sellIndia'])
                   or 'BREAKAWAY' in str(regression_data['sellIndia'])
                   or 'TRISTAR' in str(regression_data['sellIndia'])
                   or ('SHOOTINGSTAR' in str(regression_data['sellIndia']) and regression_data['PCT_day_change'] < 0)
                   or 'DARKCLOUDCOVER' in str(regression_data['sellIndia'])
                   or '3INSIDE' in str(regression_data['sellIndia'])
                   or '3OUTSIDE' in str(regression_data['sellIndia'])
                   or '2CROWS' in str(regression_data['sellIndia'])
                   or '3BLACKCROWS' in str(regression_data['sellIndia'])
                   ) and (regression_data['forecast_day_PCT5_change'] >= 5) and (regression_data['forecast_day_PCT10_change'] >= 5)):
                    ws_sellPattern2.append(regressionResult)
                    ws_sellPattern2.append(classificationResult)    
                                   
def calculateParallel(threads=2, run_type=None, futures=None):
    pool = ThreadPool(threads)
    if(run_type == 'broker'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_broker_buy.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)
    elif(run_type == 'result'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_result.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)  
    elif(run_type == 'result_declared'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_result_declared.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)               
    else:
        scrips = []
        for data in db.scrip.find({'futures':futures}):
            scrips.append(data['scrip'].replace('&','').replace('-','_'))
        scrips.sort()
        pool.map(result_data, scrips)
                     
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1, sys.argv[1], sys.argv[2])
    connection.close()
    saveReports(sys.argv[1])