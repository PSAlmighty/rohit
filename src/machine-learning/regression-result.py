import os, logging, sys, json, csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from talib.abstract import *

import datetime
import time
import gc

connection = MongoClient('localhost', 27017)
db = connection.Nsedata

directory = '../../output/final'
logname = '../../output/final' + '/regression-result' + time.strftime("%d%m%y-%H%M%S")

newsDict = {}
wb = Workbook()
ws_buyAll = wb.create_sheet("BuyAll")
ws_buyAll.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_sellAll = wb.create_sheet("SellAll")
ws_sellAll.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_buyYearHigh = wb.create_sheet("buyYearHigh")
ws_buyYearHigh.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_sellYearLow = wb.create_sheet("sellYearLow")
ws_sellYearLow.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_buyUpTrend = wb.create_sheet("buyUpTrend")
ws_buyUpTrend.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_sellDownTrend = wb.create_sheet("sellDownTrend")
ws_sellDownTrend.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_buyFinal = wb.create_sheet("BuyFinal")
ws_buyFinal.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_sellFinal = wb.create_sheet("SellFinal")
ws_sellFinal.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_buyFinal1 = wb.create_sheet("BuyFinal1")
ws_buyFinal1.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_sellFinal1 = wb.create_sheet("SellFinal1")
ws_sellFinal1.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_buyPattern = wb.create_sheet("BuyPattern")
ws_buyPattern.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_sellPattern = wb.create_sheet("SellPattern")
ws_sellPattern.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_buyPattern1 = wb.create_sheet("BuyPattern1")
ws_buyPattern1.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_sellPattern1 = wb.create_sheet("SellPattern1")
ws_sellPattern1.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_buyPattern2 = wb.create_sheet("buyPattern2")
ws_buyPattern2.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_sellPattern2 = wb.create_sheet("sellPattern2")
ws_sellPattern2.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_sellYearHigh = wb.create_sheet("sellYearHigh")
ws_sellYearHigh.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])
ws_buyYearLow = wb.create_sheet("buyYearLow")
ws_buyYearLow.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment"])

def saveReports(run_type=None):
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws_buyAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_buyAll.add_table(tab)
    
    count = 0
    for row in ws_sellAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_sellAll.add_table(tab)
    
    count = 0
    for row in ws_buyYearHigh.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_buyYearHigh.add_table(tab)
    
    count = 0
    for row in ws_sellYearHigh.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_sellYearHigh.add_table(tab)
    
    count = 0
    for row in ws_buyYearLow.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_buyYearLow.add_table(tab)
    
    count = 0
    for row in ws_sellYearLow.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_sellYearLow.add_table(tab)
    
    count = 0
    for row in ws_buyFinal.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_buyFinal.add_table(tab)
    
    count = 0
    for row in ws_sellFinal.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_sellFinal.add_table(tab)
    
    count = 0
    for row in ws_buyFinal1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_buyFinal1.add_table(tab)
    
    count = 0
    for row in ws_sellFinal1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_sellFinal1.add_table(tab)
    
    count = 0
    for row in ws_buyPattern.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_buyPattern.add_table(tab)
    
    count = 0
    for row in ws_sellPattern.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_sellPattern.add_table(tab)
      
    count = 0
    for row in ws_buyPattern1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_buyPattern1.add_table(tab)
    
    count = 0
    for row in ws_sellPattern1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_sellPattern1.add_table(tab)
    
    count = 0
    for row in ws_buyPattern2.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_buyPattern2.add_table(tab)
    
    count = 0
    for row in ws_sellPattern2.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_sellPattern2.add_table(tab)
    
    count = 0
    for row in ws_buyUpTrend.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_buyUpTrend.add_table(tab)
    
    count = 0
    for row in ws_sellDownTrend.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:X" + str(count))
    tab.tableStyleInfo = style
    ws_sellDownTrend.add_table(tab)
      
    if(run_type == 'broker'):
        wb.save(logname + "broker_buy.xlsx")
    elif(run_type == 'result'):
        wb.save(logname + "result.xlsx")
    elif(run_type == 'result_declared'):
        wb.save(logname + "result_declared.xlsx")       
    else:
        wb.save(logname + ".xlsx")

def buy_News(scrip):
    scrip_newsList = db.news.find_one({'scrip':scrip})
    ws_buyNews.append([scrip])
    ws_buyNews.append(["#####################"])
    if(scrip_newsList is None):
        print('Missing news for ', scrip)
        return
    
    for scrip_news in scrip_newsList['news']:
        ws_buyNews.append([scrip_news['timestamp'], scrip_news['summary'], scrip_news['link']])
    ws_buyNews.append([" "])
    
def sell_News(scrip):
    scrip_newsList = db.news.find_one({'scrip':scrip})
    ws_sellNews.append([scrip])
    ws_sellNews.append(["#####################"])
    if(scrip_newsList is None):
        print('Missing news for ', scrip)
        return
    for scrip_news in scrip_newsList['news']:
        ws_sellNews.append([scrip_news['timestamp'], scrip_news['summary'], scrip_news['link']])
    ws_sellNews.append([" "])  

def all_day_pct_change_negative(regression_data):
    if(regression_data['forecast_day_PCT_change'] < 0
        and regression_data['forecast_day_PCT2_change'] < .5
        and regression_data['forecast_day_PCT3_change'] < .5
        and regression_data['forecast_day_PCT4_change'] < .5
        and regression_data['forecast_day_PCT5_change'] < .5
        and regression_data['forecast_day_PCT7_change'] < .5
        and regression_data['forecast_day_PCT10_change'] < .5):
        return True;
    
def all_day_pct_change_positive(regression_data):
    if(regression_data['forecast_day_PCT_change'] > 0
        and regression_data['forecast_day_PCT2_change'] > -.5
        and regression_data['forecast_day_PCT3_change'] > -.5
        and regression_data['forecast_day_PCT4_change'] > -.5
        and regression_data['forecast_day_PCT5_change'] > -.5
        and regression_data['forecast_day_PCT7_change'] > -.5
        and regression_data['forecast_day_PCT10_change'] > -.5):
        return True;    

def result_data(scrip):
    resultDeclared = ""
    resultDate = ""
    resultSentiment = ""
    resultComment = ""
    result_data = db.scrip_result.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(result_data is not None):
        resultDate = result_data['result_date'].strip()
        resultSentiment = result_data['result_sentiment']
        resultComment = result_data['comment']
        start_date = (datetime.datetime.now() - datetime.timedelta(hours=0))
        start_date = datetime.datetime(start_date.year, start_date.month, start_date.day, start_date.hour)
        result_time = datetime.datetime.strptime(resultDate, "%Y-%m-%d")
        if result_time < start_date: 
            resultDeclared = resultDate 
            resultDate = ""
       
    regression_data = db.regressionhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None):
        regressionResult = [ ]
        regressionResult.append(regression_data['buyIndia'])
        regressionResult.append(regression_data['sellIndia'])
        regressionResult.append(regression_data['scrip'])
        regressionResult.append(regression_data['forecast_day_VOL_change'])
        regressionResult.append(regression_data['forecast_day_PCT_change'])
        regressionResult.append(regression_data['forecast_day_PCT2_change'])
        regressionResult.append(regression_data['forecast_day_PCT3_change'])
        regressionResult.append(regression_data['forecast_day_PCT4_change'])
        regressionResult.append(regression_data['forecast_day_PCT5_change'])
        regressionResult.append(regression_data['forecast_day_PCT7_change'])
        regressionResult.append(regression_data['forecast_day_PCT10_change'])
        regressionResult.append(regression_data['PCT_day_change'])
        regressionResult.append(regression_data['PCT_change'])
        regressionResult.append(regression_data['score'])
        regressionResult.append(regression_data['mlpValue'])
        regressionResult.append(regression_data['kNeighboursValue'])
        regressionResult.append(regression_data['trend'])
        regressionResult.append(regression_data['yearHighChange'])
        regressionResult.append(regression_data['yearLowChange'])
        regressionResult.append(resultDate)
        regressionResult.append(resultDeclared)
        regressionResult.append(resultSentiment)
        regressionResult.append(resultComment)
        score = ''
        if(regression_data['score'] == '10' or regression_data['score'] == '1-1'):
            score = 'up'    
        dayClose = False
        if(regression_data['PCT_day_change'] > .5 and regression_data['PCT_change'] < .1):
            dayClose = True
        regressionResult.append(dayClose)
        longTrend = False 
        if(all_day_pct_change_positive(regression_data)):
            longTrend = True
        regressionResult.append(longTrend)      
        if(((regression_data['mlpValue'] >= 1 and regression_data['kNeighboursValue'] >= 0.5) or (regression_data['mlpValue'] >= 0.5 and regression_data['kNeighboursValue'] >= 1))
           and 'P@[' not in str(regression_data['sellIndia'])):
            ws_buyAll.append(regressionResult)
            if(-5 <= regression_data['yearHighChange'] < -1 and regression_data['yearLowChange'] > 30 
                and -0.5 < regression_data['PCT_day_change'] < 3 and regression_data['forecast_day_PCT2_change'] <= 5):
                ws_buyYearHigh.append(regressionResult)
            elif(-15 < regression_data['yearHighChange'] < -5 and regression_data['yearLowChange'] > 30 
                and -0.5 < regression_data['PCT_day_change'] < 3 and (score == 'up'  or regression_data['forecast_day_PCT_change'] > 0)):
                ws_buyYearHigh.append(regressionResult)
                
            if(1 < regression_data['yearLowChange'] < 10 and regression_data['yearHighChange'] < -30 
                and 0 < regression_data['PCT_day_change'] < 3 and score == 'up'
                and (regression_data['forecast_day_PCT10_change'] <= -5 or regression_data['forecast_day_PCT5_change'] > 5)):
                ws_buyYearLow.append(regressionResult) 
                
            if(longTrend and 0 < regression_data['PCT_day_change'] < 8 and regression_data['yearHighChange'] < -10
                and regression_data['forecast_day_PCT10_change'] >= regression_data['PCT_change'] + 2
                and 'P@[' in str(regression_data['buyIndia'])):
                ws_buyUpTrend.append(regressionResult)    
                   
            if(regression_data['yearHighChange'] < -5 and regression_data['yearLowChange'] > 5 and regression_data['score'] != '0-1'):   
                if(2 > regression_data['PCT_day_change'] > 0 and str(regression_data['sellIndia']) == '' and -95 < regression_data['yearHighChange'] < -15
                    and regression_data['forecast_day_PCT5_change'] <= 1 and regression_data['forecast_day_PCT7_change'] <= -1 and regression_data['forecast_day_PCT10_change'] <= -5):
                    ws_buyFinal.append(regressionResult) 
                elif(2 > regression_data['PCT_day_change'] > 0
                    and regression_data['forecast_day_PCT5_change'] <= 1 and regression_data['forecast_day_PCT7_change'] <= -1 and regression_data['forecast_day_PCT10_change'] <= -5):
                    ws_buyFinal1.append(regressionResult) 
                 
            if(regression_data['PCT_day_change'] < 4 and regression_data['yearLowChange'] > 5):
                if(('MARUBOZU' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and regression_data['forecast_day_PCT10_change'] <= 1)
                   or ('HAMMER' in str(regression_data['buyIndia']) and regression_data['PCT_day_change'] > 0)
                   #or 'ENGULFING' in str(regression_data['buyIndia'])
                   #or 'PIERCING' in str(regression_data['buyIndia'])
                   or ('MORNINGSTAR' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and regression_data['forecast_day_PCT10_change'] <= 1)
                   #or ':DOJISTAR' in str(regression_data['buyIndia'])
                   #or 'MORNINGDOJISTAR' in str(regression_data['buyIndia'])
                   or 'ABANDONEDBABY' in str(regression_data['buyIndia'])
                   or 'COUNTERATTACK' in str(regression_data['buyIndia'])
                   or 'KICKING' in str(regression_data['buyIndia'])
                   or 'BREAKAWAY' in str(regression_data['buyIndia'])
                   #or 'TRISTAR' in str(regression_data['buyIndia'])
                   #or '3WHITESOLDIERS' in str(regression_data['buyIndia'])
                   #or '3INSIDE' in str(regression_data['buyIndia'])
                   ):
                    ws_buyPattern.append(regressionResult) 
                elif(('CCI:BOP' in str(regression_data['buyIndia']) and 'BELTHOLD' in str(regression_data['buyIndia']))
                   or ('AROON:BOP' in str(regression_data['buyIndia']) and 'BELTHOLD' in str(regression_data['buyIndia']) and 'ENGULFING' in str(regression_data['buyIndia']))
                   or ('BELTHOLD' == str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and score == 'up')
                   #or ('3OUTSIDE' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and score == 'up')
                   #or ('HARAMI' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and score == 'up')
                   #or (regression_data['yearHighChange'] <= -35 and 'HARAMI' in str(regression_data['buyIndia']) and 'SHORTLINE' in str(regression_data['buyIndia']) and regression_data['PCT_day_change'] > 0)
                   or ('DOJI' in str(regression_data['buyIndia']) and 'GRAVESTONEDOJI' in str(regression_data['buyIndia']) and 'LONGLEGGEDDOJI' in str(regression_data['buyIndia']) and regression_data['PCT_day_change'] > 0)
                   #or ('P@[,HIKKAKE]' == str(regression_data['buyIndia']) and regression_data['PCT_day_change'] < 0)
                   #or (regression_data['yearHighChange'] <= -35 and 'BELTHOLD' in str(regression_data['buyIndia']) and 'LONGLINE' in str(regression_data['buyIndia']))
                   #or (regression_data['yearHighChange'] <= -35 and ',CCI:BOP' in str(regression_data['buyIndia']) and 'LONGLINE' in str(regression_data['buyIndia']))
                   ):
                    ws_buyPattern1.append(regressionResult)
                elif((('MARUBOZU' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and regression_data['forecast_day_PCT10_change'] <= 1)
                   or ('HAMMER' in str(regression_data['buyIndia']) and regression_data['PCT_day_change'] > 0)
                   or 'ENGULFING' in str(regression_data['buyIndia'])
                   or 'PIERCING' in str(regression_data['buyIndia'])
                   or ('MORNINGSTAR' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and regression_data['forecast_day_PCT10_change'] <= 1)
                   #or ':DOJISTAR' in str(regression_data['buyIndia'])
                   or 'MORNINGDOJISTAR' in str(regression_data['buyIndia'])
                   or 'ABANDONEDBABY' in str(regression_data['buyIndia'])
                   or 'COUNTERATTACK' in str(regression_data['buyIndia'])
                   or 'KICKING' in str(regression_data['buyIndia'])
                   or 'BREAKAWAY' in str(regression_data['buyIndia'])
                   or 'TRISTAR' in str(regression_data['buyIndia'])
                   or '3WHITESOLDIERS' in str(regression_data['buyIndia'])
                   or '3INSIDE' in str(regression_data['buyIndia'])
                   ) 
                    and 'DOJI' not in str(regression_data['buyIndia']) 
                    and (regression_data['forecast_day_PCT5_change'] <= -5) and (regression_data['forecast_day_PCT10_change'] <= -5)): 
                    ws_buyPattern2.append(regressionResult)
                elif('BELTHOLD, LONGLINE' in str(regression_data['buyIndia'])):
                    ws_buyPattern2.append(regressionResult)
                    
    regression_data = db.regressionlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None):
        regressionResult = [ ]
        regressionResult.append(regression_data['buyIndia'])
        regressionResult.append(regression_data['sellIndia'])
        regressionResult.append(regression_data['scrip'])
        regressionResult.append(regression_data['forecast_day_VOL_change'])
        regressionResult.append(regression_data['forecast_day_PCT_change'])
        regressionResult.append(regression_data['forecast_day_PCT2_change'])
        regressionResult.append(regression_data['forecast_day_PCT3_change'])
        regressionResult.append(regression_data['forecast_day_PCT4_change'])
        regressionResult.append(regression_data['forecast_day_PCT5_change'])
        regressionResult.append(regression_data['forecast_day_PCT7_change'])
        regressionResult.append(regression_data['forecast_day_PCT10_change'])
        regressionResult.append(regression_data['PCT_day_change'])
        regressionResult.append(regression_data['PCT_change'])
        regressionResult.append(regression_data['score'])
        regressionResult.append(regression_data['mlpValue'])
        regressionResult.append(regression_data['kNeighboursValue'])
        regressionResult.append(regression_data['trend'])
        regressionResult.append(regression_data['yearHighChange'])
        regressionResult.append(regression_data['yearLowChange'])
        regressionResult.append(resultDate)
        regressionResult.append(resultDeclared)
        regressionResult.append(resultSentiment)
        regressionResult.append(resultComment)
        score = ''
        if(regression_data['score'] == '1-1' or regression_data['score'] == '0-1'):
            score = 'down'
        dayClose = False
        if(regression_data['PCT_day_change'] < -.5 and regression_data['PCT_change'] > -.1):
            dayClose = True   
        regressionResult.append(dayClose)
        longTrend = False 
        if(all_day_pct_change_negative(regression_data)):
            longTrend = True
        regressionResult.append(longTrend)       
        if(((regression_data['mlpValue'] <= -1 and regression_data['kNeighboursValue'] <= -0.5) or (regression_data['mlpValue'] <= -0.5 and regression_data['kNeighboursValue'] <= -1))
           and 'P@[' not in str(regression_data['buyIndia'])):
            ws_sellAll.append(regressionResult)
            if(-10 < regression_data['yearHighChange'] < -2 and regression_data['yearLowChange'] > 30 and -2 < regression_data['PCT_day_change'] < 0.8 
                and regression_data['forecast_day_PCT5_change'] > 0 and regression_data['forecast_day_PCT10_change'] > 5 
                and (score == 'down'  or regression_data['forecast_day_PCT_change'] < 0)):
                ws_sellYearHigh.append(regressionResult)
                
            if(0 < regression_data['yearLowChange'] < 15 and regression_data['yearHighChange'] < -30 
                and -2 < regression_data['PCT_day_change'] < 0 and regression_data['forecast_day_PCT_change'] < 0
                and all_day_pct_change_negative(regression_data)):
                ws_sellYearLow.append(regressionResult)
                
            if(longTrend and -8 < regression_data['PCT_day_change'] < 0 and regression_data['yearLowChange'] > 10
                and regression_data['forecast_day_PCT10_change'] <= regression_data['PCT_change'] - 2):
                ws_sellDownTrend.append(regressionResult)      
            
            if(regression_data['yearHighChange'] < -5 and regression_data['yearLowChange'] > 5 and regression_data['score'] != '10'):
                if(-2 < regression_data['PCT_day_change'] < 0 and str(regression_data['buyIndia']) == '' and -95 < regression_data['yearHighChange'] < -20
                    and regression_data['forecast_day_PCT5_change'] >= 1 and regression_data['forecast_day_PCT7_change'] >= 1 and regression_data['forecast_day_PCT10_change'] >= 5):
                    ws_sellFinal.append(regressionResult)
                elif(-2 < regression_data['PCT_day_change'] < 0 
                    and regression_data['forecast_day_PCT5_change'] >= -1 and regression_data['forecast_day_PCT7_change'] >= 1 and regression_data['forecast_day_PCT10_change'] >= 5):
                    ws_sellFinal1.append(regressionResult)    
           
        
            if(-4 < regression_data['PCT_day_change'] and regression_data['yearHighChange'] < -5):
                if(('HANGINGMAN' in str(regression_data['sellIndia'])
                   #or 'MARUBOZU' in str(regression_data['sellIndia'])
                   #or 'ENGULFING' in str(regression_data['sellIndia'])
                   or 'EVENINGSTAR' in str(regression_data['sellIndia'])
                   #or ':DOJISTAR' in str(regression_data['sellIndia'])
                   #or 'EVENINGDOJISTAR' in str(regression_data['sellIndia'])
                   or 'ABANDONEDBABY' in str(regression_data['sellIndia'])
                   or 'COUNTERATTACK' in str(regression_data['sellIndia'])
                   or 'KICKING' in str(regression_data['sellIndia'])
                   or 'BREAKAWAY' in str(regression_data['sellIndia'])
                   #or 'TRISTAR' in str(regression_data['sellIndia'])
                   or ('SHOOTINGSTAR' in str(regression_data['sellIndia']) and regression_data['PCT_day_change'] < 0)
                   or 'DARKCLOUDCOVER' in str(regression_data['sellIndia'])
                   #or '3INSIDE' in str(regression_data['sellIndia'])
                   #or '3OUTSIDE' in str(regression_data['sellIndia'])
                   #or '2CROWS' in str(regression_data['sellIndia'])
                   #or '3BLACKCROWS' in str(regression_data['sellIndia'])
                   ) and (regression_data['forecast_day_PCT5_change'] >= 0)):
                    ws_sellPattern.append(regressionResult)
                elif((('HARAMI' in str(regression_data['sellIndia']) and regression_data['forecast_day_PCT5_change'] >= 0 and score == 'down')
                   or ('ENGULFING' in str(regression_data['sellIndia']) and 'LONGLINE' in str(regression_data['sellIndia']) and score == 'down')
                   ) and regression_data['yearHighChange'] < -5):
                    ws_sellPattern1.append(regressionResult)
                elif(('HANGINGMAN' in str(regression_data['sellIndia'])
                   or 'MARUBOZU' in str(regression_data['sellIndia'])
                   #or 'ENGULFING' in str(regression_data['sellIndia'])
                   or 'EVENINGSTAR' in str(regression_data['sellIndia'])
                   #or ':DOJISTAR' in str(regression_data['sellIndia'])
                   #or 'EVENINGDOJISTAR' in str(regression_data['sellIndia'])
                   or 'ABANDONEDBABY' in str(regression_data['sellIndia'])
                   or 'COUNTERATTACK' in str(regression_data['sellIndia'])
                   or 'KICKING' in str(regression_data['sellIndia'])
                   or 'BREAKAWAY' in str(regression_data['sellIndia'])
                   or 'TRISTAR' in str(regression_data['sellIndia'])
                   or ('SHOOTINGSTAR' in str(regression_data['sellIndia']) and regression_data['PCT_day_change'] < 0)
                   or 'DARKCLOUDCOVER' in str(regression_data['sellIndia'])
                   #or '3INSIDE' in str(regression_data['sellIndia'])
                   #or '3OUTSIDE' in str(regression_data['sellIndia'])
                   or '2CROWS' in str(regression_data['sellIndia'])
                   or '3BLACKCROWS' in str(regression_data['sellIndia'])
                   ) and (regression_data['forecast_day_PCT5_change'] >= 5) and (regression_data['forecast_day_PCT10_change'] >= 5)):
                    ws_sellPattern2.append(regressionResult)                                 
                elif(('BELTHOLD' in str(regression_data['sellIndia']) and 'LONGLINE' in str(regression_data['sellIndia']) and score == '0-1')
                     or ('CLOSINGMARUBOZU' in str(regression_data['sellIndia']) and 'LONGLINE' in str(regression_data['sellIndia']))
                     or ('M@[,CROSSOVER-MACD]' in str(regression_data['sellIndia']) and 'LONGLINE' in str(regression_data['sellIndia']))
                     or ('3OUTSIDE' in str(regression_data['sellIndia']) and 'SPINNINGTOP' not in str(regression_data['sellIndia']) and 'P@[3OUTSIDE]' != str(regression_data['sellIndia']))
                     ):
                    ws_sellPattern2.append(regressionResult)
            
def calculateParallel(threads=2, run_type=None, futures=None):
    pool = ThreadPool(threads)
    if(run_type == 'broker'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_broker_buy.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)
    elif(run_type == 'result'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_result.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)  
    elif(run_type == 'result_declared'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_result_declared.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)               
    else:
        scrips = []
        for data in db.scrip.find({'futures':futures}):
            scrips.append(data['scrip'].replace('&','').replace('-','_'))
        scrips.sort()
        pool.map(result_data, scrips)
                     
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1, sys.argv[1], sys.argv[2])
    connection.close()
    saveReports(sys.argv[1])